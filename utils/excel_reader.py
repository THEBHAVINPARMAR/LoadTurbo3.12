import openpyxl

def get_login_data(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping header row
        if all(row):  # Skip rows with any empty cell
            data.append({
                "email": row[0],
                "password": row[1]
            })

    wb.close()
    return data

def get_register_data(excel_path):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping header row
        if all(row):  # Skip rows with any empty cell
            data.append({
                "first_name": row[0],
                "last_name": row[1],
                "mobile": row[2],
                "email": row[3],
                "access_code": row[4],
                "password": row[5],
                "confirm_password": row[6]
            })

    wb.close()
    return data

